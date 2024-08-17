import pandas as pd
from difflib import SequenceMatcher
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def preparar_dataframes(contactos_path, egresados_path):
    contactos_df = pd.read_excel(contactos_path)
    egresados_df = pd.read_excel(egresados_path)
    contactos_df['Nombre'] = contactos_df[['First Name', 'Middle Name', 'Last Name']].fillna('').agg(' '.join, axis=1).str.strip()
    return contactos_df, egresados_df

def remove_accents(input_str):
    replacements = (
        ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"),
        ("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U")
    )
    for a, b in replacements:
        input_str = input_str.replace(a, b)
    return input_str

def normalize_name(name):
    name = str(name).strip()
    name = remove_accents(name)
    name = name.upper()
    return ' '.join(name.split())

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def calcular_coincidencias(nombre1, nombre2):
    palabras1 = set(nombre1.split())
    palabras2 = set(nombre2.split())
    coincidencias = palabras1.intersection(palabras2)
    return len(coincidencias)

def ajustar_filas_y_columnas(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    for row in worksheet.iter_rows():
        max_height = 0
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.value:
                # Ensure the value is a string before using count()
                cell_text = str(cell.value)
                cell_height = cell_text.count('\n') + 1
                if cell_height > max_height:
                    max_height = cell_height
        worksheet.row_dimensions[cell.row].height = max_height * 15  # Approximate height per line

def generar_archivo_combinado(contactos_path, egresados_path, output_path, progress_bar):
    contactos_df, egresados_df = preparar_dataframes(contactos_path, egresados_path)
    contactos_df['Nombre'] = contactos_df['Nombre'].apply(normalize_name)
    egresados_df['Nombres'] = egresados_df['Nombres'].apply(normalize_name)

    keyword_index = defaultdict(list)
    for index, egresado in egresados_df.iterrows():
        words = set(egresado['Nombres'].split())
        for word in words:
            keyword_index[word].append((index, egresado['Nombres']))

    resultados = []

    total_rows = len(contactos_df)
    for idx, contacto in enumerate(contactos_df.iterrows()):
        index, contacto = contacto
        nombre_contacto = contacto['Nombre']
        palabras_contacto = set(nombre_contacto.split())
        posibles_candidatos = set()
        for palabra in palabras_contacto:
            if palabra in keyword_index:
                posibles_candidatos.update(keyword_index[palabra])
        mejor_match = None
        mejor_similitud = 0
        for candidato in posibles_candidatos:
            index2, nombre_egresado = candidato
            similitud = similar(nombre_contacto, nombre_egresado)
            if similitud > mejor_similitud:
                mejor_similitud = similitud
                mejor_match = egresados_df.loc[index2]
        if mejor_match is not None:
            resultados.append({
                'Cedula': mejor_match['Cedula'],
                'Nombre': nombre_contacto,
                'Telefono1': contacto['Phone 1 - Value'],
                'Telefono2': contacto['Phone 2 - Value'],
                'Nombre Egresado': mejor_match['Nombres'],
                'Certeza': mejor_similitud * 100,
                'Coincidencias': calcular_coincidencias(nombre_contacto, mejor_match['Nombres'])
            })

        # Update progress
        progress_bar.progress(int((idx + 1) / total_rows * 100))

    # Convertir los resultados en un DataFrame
    resultados_df = pd.DataFrame(resultados)

    # Ordenar por Coincidencias y luego por Certeza
    resultados_df.sort_values(by=['Coincidencias', 'Certeza'], ascending=[False, False], inplace=True)

    # Guardar el resultado en un archivo Excel
    resultados_df.to_excel(output_path, index=False)

    # Ajustar las filas y columnas para que ocupen el espacio necesario
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    ajustar_filas_y_columnas(worksheet)
    workbook.save(output_path)
