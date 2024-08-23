import pandas as pd
from difflib import SequenceMatcher
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from io import BytesIO

def preparar_dataframes(contactos_data, egresados_data):
    contactos_df = pd.read_excel(contactos_data)
    egresados_df = pd.read_excel(egresados_data)
    contactos_df['Nombre'] = contactos_df[['First Name', 'Middle Name', 'Last Name']].fillna('').agg(' '.join, axis=1).str.strip()
    return contactos_df, egresados_df

def remove_accents(input_str):
    replacements = (
        ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"),
        ("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U")
    )
    for a, b in replacements:
        input_str = input_str.replace(a, b)
    return input_str

def normalize_name(name):
    name = str(name).strip()
    name = remove_accents(name)
    name = name.upper()
    return ' '.join(name.split())

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def calcular_coincidencias(nombre1, nombre2):
    palabras1 = set(nombre1.split())
    palabras2 = set(nombre2.split())
    coincidencias = palabras1.intersection(palabras2)
    return len(coincidencias)

def ajustar_filas_y_columnas(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception as e:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    for row in worksheet.iter_rows():
        max_height = 0
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.value:
                cell_text = str(cell.value)
                cell_height = cell_text.count('\n') + 1
                if cell_height > max_height:
                    max_height = cell_height
        worksheet.row_dimensions[cell.row].height = max_height * 15  # Approximate height per line

def limpiar_telefono(telefono):
    if pd.isna(telefono):
        return telefono
    telefono = telefono.replace(" ", "")
    if telefono.startswith("+"):
        return telefono
    if telefono.startswith("57"):
        return f"+{telefono}"
    if telefono.startswith("3"):
        return f"+57{telefono}"
    return telefono

def extraer_primer_nombre(nombre):
    return nombre.split()[0].capitalize()

def generar_archivo_combinado(contactos_data, egresados_data, output_stream, progress_bar):
    contactos_df, egresados_df = preparar_dataframes(contactos_data, egresados_data)
    contactos_df['Nombre'] = contactos_df['Nombre'].apply(normalize_name)
    egresados_df['Nombres'] = egresados_df['Nombres'].apply(normalize_name)

    telefono1_col = 'Phone 1 - Value' if 'Phone 1 - Value' in contactos_df.columns else 'Mobile Phone'
    telefono2_col = 'Phone 2 - Value' if 'Phone 2 - Value' in contactos_df.columns else ('Other Phone' if 'Other Phone' in contactos_df.columns else None)
    
    # Asegurar que exista la columna Telefono 2, incluso si no está en los datos originales
    if telefono2_col not in contactos_df.columns:
        contactos_df[telefono2_col] = ''

    keyword_index = defaultdict(list)
    for index, egresado in egresados_df.iterrows():
        words = set(egresado['Nombres'].split())
        for word in words:
            keyword_index[word].append((index, egresado['Nombres']))

    resultados = []

    total_rows = len(contactos_df)
    for idx, contacto in enumerate(contactos_df.iterrows()):
        index, contacto = contacto
        nombre_contacto = contacto['Nombre']
        palabras_contacto = set(nombre_contacto.split())
        posibles_candidatos = set()
        for palabra in palabras_contacto:
            if palabra in keyword_index:
                posibles_candidatos.update(keyword_index[palabra])
        mejor_match = None
        mejor_similitud = 0
        mejor_coincidencias = 0
        for candidato in posibles_candidatos:
            index2, nombre_egresado = candidato
            similitud = similar(nombre_contacto, nombre_egresado)
            coincidencias = calcular_coincidencias(nombre_contacto, nombre_egresado)
            if coincidencias > mejor_coincidencias or (coincidencias == mejor_coincidencias and similitud > mejor_similitud):
                mejor_coincidencias = coincidencias
                mejor_similitud = similitud
                mejor_match = egresados_df.loc[index2]
        if mejor_match is not None:
            porcentaje_coincidencias = (mejor_coincidencias / len(set(mejor_match['Nombres'].split()))) * 100
            promedio_acierto = (mejor_similitud * 100 + porcentaje_coincidencias) / 2
            resultado = {
                'Cedula': mejor_match['Cedula'],
                'Mi Contacto': nombre_contacto,
                'Encontrado': mejor_match['Nombres'],  # Cambio de "Nombre votante" a "Encontrado"
                'Tipo': mejor_match['Tipo'],  # Tipo incluido
                'Telefono1': contacto[telefono1_col],
                'Telefono2': contacto[telefono2_col],  # Añadido Telefono2
                'Acierto [%]': promedio_acierto,
                'TELEFONO': limpiar_telefono(contacto[telefono1_col]) if not pd.isna(contacto[telefono1_col]) else limpiar_telefono(contacto[telefono2_col]),
                'PrimerNombre': extraer_primer_nombre(nombre_contacto)
            }

            resultados.append(resultado)

        progress_bar.progress(int((idx + 1) / total_rows * 100))

    resultados_df = pd.DataFrame(resultados)

    resultados_df.sort_values(by=['Acierto [%]'], ascending=False, inplace=True)

    workbook = Workbook()
    worksheet = workbook.active

    for r_idx, row in enumerate(dataframe_to_rows(resultados_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    ajustar_filas_y_columnas(worksheet)

    workbook.save(output_stream)
    output_stream.seek(0)

def generar_archivo_filtro_unillanos(contactos_data, egresados_data, output_stream, progress_bar):
    contactos_df, egresados_df = preparar_dataframes(contactos_data, egresados_data)
    contactos_df['Nombre'] = contactos_df['Nombre'].apply(normalize_name)
    egresados_df['Nombres'] = egresados_df['Nombres'].apply(normalize_name)

    telefono1_col = 'Phone 1 - Value' if 'Phone 1 - Value' in contactos_df.columns else 'Mobile Phone'
    telefono2_col = 'Phone 2 - Value' if 'Phone 2 - Value' in contactos_df.columns else ('Other Phone' if 'Other Phone' in contactos_df.columns else None)

    if telefono2_col not in contactos_df.columns:
        contactos_df[telefono2_col] = ''

    keyword_index = defaultdict(list)
    for index, egresado in egresados_df.iterrows():
        words = set(egresado['Nombres'].split())
        for word in words:
            keyword_index[word].append((index, egresado['Nombres']))

    resultados = []

    total_rows = len(contactos_df)
    for idx, contacto in enumerate(contactos_df.iterrows()):
        index, contacto = contacto
        nombre_contacto = contacto['Nombre']
        palabras_contacto = set(nombre_contacto.split())
        posibles_candidatos = set()
        for palabra in palabras_contacto:
            if palabra in keyword_index:
                posibles_candidatos.update(keyword_index[palabra])
        mejor_match = None
        mejor_similitud = 0
        mejor_coincidencias = 0
        for candidato in posibles_candidatos:
            index2, nombre_egresado = candidato
            similitud = similar(nombre_contacto, nombre_egresado)
            coincidencias = calcular_coincidencias(nombre_contacto, nombre_egresado)
            if coincidencias > mejor_coincidencias or (coincidencias == mejor_coincidencias and similitud > mejor_similitud):
                mejor_coincidencias = coincidencias
                mejor_similitud = similitud
                mejor_match = egresados_df.loc[index2]
        if mejor_match is not None:
            porcentaje_coincidencias = (mejor_coincidencias / len(set(mejor_match['Nombres'].split()))) * 100
            promedio_acierto = (mejor_similitud * 100 + porcentaje_coincidencias) / 2
            resultado = {
                'Cedula': mejor_match['Cedula'],
                'Mi Contacto': nombre_contacto,
                'Encontrado': mejor_match['Nombres'],
                'Tipo': mejor_match['Tipo'],
                'Telefono1': contacto[telefono1_col],
                'Telefono2': contacto[telefono2_col],
                'Acierto [%]': promedio_acierto,
                'TELEFONO': limpiar_telefono(contacto[telefono1_col]) if not pd.isna(contacto[telefono1_col]) else limpiar_telefono(contacto[telefono2_col]),
                'PrimerNombre': extraer_primer_nombre(nombre_contacto)
            }

            resultados.append(resultado)

        progress_bar.progress(int((idx + 1) / total_rows * 100))

    resultados_df = pd.DataFrame(resultados)

    # Ordenar por 'Acierto [%]' de mayor a menor
    resultados_df.sort_values(by=['Acierto [%]'], ascending=False, inplace=True)

    # Filtrar y ordenar primero los contactos que contienen 'U' o 'Unillanos' (no case sensitive)
    mask_u_unillanos = resultados_df['Mi Contacto'].str.contains(r'\bU\b|\bUnillanos\b', case=False, regex=True)
    resultados_df = pd.concat([resultados_df[mask_u_unillanos], resultados_df[~mask_u_unillanos]])

    workbook = Workbook()
    worksheet = workbook.active

    for r_idx, row in enumerate(dataframe_to_rows(resultados_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    ajustar_filas_y_columnas(worksheet)

    workbook.save(output_stream)
    output_stream.seek(0)